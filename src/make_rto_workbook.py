#!/usr/bin/env python
"""
Build the RTO-facing "Observed Ground-Truth" workbook for the dashboard.

One Excel the Kashmir RTO can open and act on — everything the app GPS layer
has established, with the honesty caveats ON the sheets (not in a footnote
somewhere else):

  1. Read Me            — what this is, tiers, and the partial-adoption caveat
  2. Observed Stops     — Tier-1 + Tier-2, coded in the plan's district-sector
                          terminology (X-series = observed, provisional)
  3. Verified Corridors — the 18 corridor verdicts (post-audit)
  4. Route Evidence     — all 186 plan routes: % of road alignment driven
  5. Local Connectors   — the 2 unmatched observed connectors (thin evidence)

Output: dashboard public/route-rationalization-kashmir/Kashmir_Observed_GroundTruth_v1.xlsx
        + a copy in data/ for the repo record.

Run:  & "D:\\plotting\\ana\\python.exe" src\\make_rto_workbook.py
"""
import os, sys, json, glob
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from common import DATA

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DASH = os.environ.get("DASH_REPO", "E:/dash/bus-sathi-dashboard")
OUT = os.path.join(DASH, "public", "route-rationalization-kashmir", "Kashmir_Observed_GroundTruth_v1.xlsx")

INK = "102A2E"; TEAL = "0F6E56"; MIST = "E9F2EE"; AMBER = "D97706"; WHITE = "FFFFFF"
HDR_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=MIST)
HDR_FONT = Font(bold=True, color=WHITE, size=11, name="Calibri")
TITLE_FONT = Font(bold=True, size=15, color=INK, name="Cambria")
NOTE_FONT = Font(italic=True, size=10, color="5B6B6B", name="Calibri")
THIN = Border(*[Side(style="thin", color="D9E2DE")] * 4)


def style_table(ws, df, start_row=3, widths=None):
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(start_row, j, col)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(i, j, v)
            c.border = THIN
            c.font = Font(size=10, name="Calibri")
            if (i - start_row) % 2 == 0:
                c.fill = ALT_FILL
    for j, col in enumerate(df.columns, start=1):
        w = (widths or {}).get(col)
        ws.column_dimensions[get_column_letter(j)].width = w or max(10, min(38, int(df[col].astype(str).str.len().quantile(0.9)) + 4))
    ws.freeze_panes = ws.cell(start_row + 1, 1)


def sheet_title(ws, title, note):
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(2, 1, note).font = NOTE_FONT


def main():
    wb = Workbook(); wb.remove(wb.active)

    # ── 1. Read Me ──────────────────────────────────────────────────
    ws = wb.create_sheet("Read Me")
    lines = [
        ("Kashmir Observed Ground-Truth — from the Bus Sathi app's real driver GPS", TITLE_FONT),
        ("", None),
        ("WHAT THIS IS", Font(bold=True, size=12, color=TEAL, name="Cambria")),
        ("Every layer in this workbook is MEASURED from ~157 working bus drivers' GPS traces "
         "(2,526 cleaned service runs, Feb–Jun 2026) — not modelled, not assumed.", None),
        ("", None),
        ("HOW TO READ THE TIERS", Font(bold=True, size=12, color=TEAL, name="Cambria")),
        ("Tier 1 = strong repeated evidence (multiple distinct drivers, many visits/runs).", None),
        ("Tier 2 = real repeated pattern on thinner evidence — for FIELD VALIDATION, not publication.", None),
        ("X-series stop codes (e.g. SR-03-X01) use the plan's district-sector terminology but are "
         "OBSERVED stops, provisional until surveyed — they do not collide with the official register.", None),
        ("", None),
        ("THE ONE CAVEAT THAT GOVERNS EVERYTHING", Font(bold=True, size=12, color=AMBER, name="Cambria")),
        ("App adoption is partial and Srinagar-concentrated. This data validates geometry, stops and "
         "speeds. It does NOT measure ridership, demand, or service frequency — absence of app data "
         "on a route never means absence of service.", None),
        ("", None),
        ("Sheets: Observed Stops · Verified Corridors · Route Evidence · Local Connectors", NOTE_FONT),
        ("Source: github.com/Princu-Babu/bus-sathi-trace-intelligence (method, audit log, reality checks)", NOTE_FONT),
    ]
    for i, (txt, font) in enumerate(lines, start=1):
        c = ws.cell(i, 1, txt)
        if font: c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110

    # ── 2. Observed stops ───────────────────────────────────────────
    st = pd.read_csv(os.path.join(DATA, "observed_stops_coded.csv"))
    st = st.rename(columns={"stop_code": "Stop Code (X-series)", "tier": "Tier", "place": "Place (OSM)",
                            "district": "District", "tehsil": "Tehsil (Sector)", "sector": "Sector #",
                            "visits": "Observed visits", "drivers": "Distinct drivers",
                            "dwell_s": "Median dwell (s)", "lat": "Lat", "lon": "Lon",
                            "nearest_master_stop": "Nearest register stop",
                            "nearest_master_code": "Register code", "nearest_master_m": "Distance (m)"})
    ws = wb.create_sheet("Observed Stops")
    sheet_title(ws, f"Observed stops — {len(st)} coded in plan terminology",
                "X-series = observed via GPS, provisional. Tier 2 = field-validate before use. "
                "'Nearest register stop' anchors each one to the official register.")
    style_table(ws, st, widths={"Place (OSM)": 24, "Nearest register stop": 22})

    # ── 3. Verified corridors ───────────────────────────────────────
    vj = json.load(open(os.path.join(ROOT, "analyst", "corridors_verdicts.geojson"), encoding="utf-8"))
    cor = pd.DataFrame([{
        "Corridor": f"C{f['properties']['corridor_id']}",
        "Observed O–D": f["properties"].get("od", ""),
        "Verdict": f["properties"].get("class", ""),
        "Matched plan route": f["properties"].get("matched_permit", "") or "—",
        "Confidence": f["properties"].get("confidence", ""),
        "Runs": f["properties"].get("n_runs", 0),
        "Drivers": f["properties"].get("n_drivers", 0),
        "Median km": f["properties"].get("median_km", 0),
    } for f in vj["features"]]).sort_values(["Verdict", "Runs"], ascending=[True, False])
    ws = wb.create_sheet("Verified Corridors")
    sheet_title(ws, "Observed corridors vs the plan — 18 individually judged",
                "matched = same corridor/terminals · partial = geometry diverges (reconciliation list) · "
                "out_of_area = Jammu division · artifact = excluded. 0 informal/unpermitted (raw-permit checked).")
    style_table(ws, cor, widths={"Observed O–D": 34, "Matched plan route": 28})

    # ── 4. Route evidence ───────────────────────────────────────────
    ev = pd.read_csv(os.path.join(DATA, "route_evidence.csv"))
    ev["Evidence"] = ev.apply(lambda r: "STRONG" if r.obs_frac >= 0.5 and r.n_drivers >= 2
                              else ("PARTIAL" if r.obs_frac >= 0.2 and r.n_drivers >= 2 else "LITTLE"), axis=1)
    ev = ev.rename(columns={"route_id": "Plan route", "route_name": "Route name", "route_type": "Type",
                            "km": "Route km", "obs_frac": "Share of road driven", "obs_km": "Km driven",
                            "n_runs": "Run fragments", "n_drivers": "Distinct drivers"})
    ws = wb.create_sheet("Route Evidence")
    sheet_title(ws, "Per-route road-driven evidence — all 186 plan routes",
                "Fragment aggregation: broken GPS runs vote for the road-km they cover. ROAD-LEVEL evidence "
                "('buses drive this alignment'), NOT proof a specific permit operates end-to-end — "
                "parallel routes sharing a road share its evidence.")
    style_table(ws, ev, widths={"Route name": 34})

    # ── 5. Local connectors ─────────────────────────────────────────
    conns = []
    for p in sorted(glob.glob(os.path.join(ROOT, "analyst", "verdicts_tier2", "T*.json"))):
        v = json.load(open(p, encoding="utf-8"))
        conns.append({"ID": f"T{v['tail_id']}", "Observed O–D": v["od_description"],
                      "Runs": v["n_runs"], "Drivers": v["n_drivers"], "Km": v["median_km"],
                      "Assessment": v["recommendation"]})
    ws = wb.create_sheet("Local Connectors")
    sheet_title(ws, "Unmatched observed local connectors — 2 (thin evidence)",
                "Repeated observed paths matching no plan route AND no raw permit by name. Both endpoint "
                "areas are otherwise well-permitted; both are below the evidence bar to act on — listed "
                "for awareness and future field validation only.")
    style_table(ws, pd.DataFrame(conns), widths={"Observed O–D": 30, "Assessment": 70})

    wb.save(OUT)
    import shutil
    shutil.copy2(OUT, os.path.join(DATA, "Kashmir_Observed_GroundTruth_v1.xlsx"))
    print(f"Wrote {OUT}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
